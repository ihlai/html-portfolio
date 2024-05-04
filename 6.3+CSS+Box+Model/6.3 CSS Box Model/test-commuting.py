#pip install selenium
#pip install openpyxl
#pip install bs4
#download latedst chromedriver and put in the same folder
#update file name
#update date for which the fee chage becomes effective
#import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import openpyxl
import bs4
import urllib.request as req

import time

def driver():
    # 設定chrome driver的執行檔路徑
    options=Options()
    options.chrome_executable_path = r"C:\Users\ihsin\Documents\Python\chromedriver.exe"
    #建立Driver物件實體 用程式操作瀏覽器運作
    driver=webdriver.Chrome(options=options)
    #連線到Leetcode工作搜尋網頁
    # driver.get('https://www.jorudan.co.jp/norikae/teiki.html')
    # driver.get("https://www.jorudan.co.jp/norikae/cgi/nori.cgi?eki1="+"梶が谷"+"&eki2="+"二子玉川"+"&via_on=1&eki3="+""+"&eki4="+""+"&eki5="+""+"&eki6=&Dyy=2023&Dmm=10&Ddd=1&Dhh=15&Dmn1=0&Dmn2=5&Cfp=1&Czu=2&C7=1&C2=0&C3=0&C1=0&sort=rec&C4=5&C5=0&C6=2&S=%E6%A4%9C%E7%B4%A2&Cmap1=&rf=nr&pg=20&eok1=&eok2=&eok3=&eok4=&eok5=&eok6=&Csg=1")
    return driver


# wb = load_workbook("Route.xlsx")
# ws = wb.active
# s1 = wb['Sheet1']
def get_values(sheet):
    list_stations = []
    r = sheet.max_row
    c= sheet.max_column
    for y in range(r):
        list_stations.append([])
        for x in range(c):
            value = sheet.cell(y+1, x+1).value
            if value == None:
                value =""
            list_stations[y].append(value)
    return list_stations
# print(get_values(s1))

# # locate fields and fill with data
# for index, row in excel_file.iterrows():
def fill_in_data(list_stations, nowrow):
    field1 = driver.find_element(By.ID,"eki1_in")
    field1.send_keys(list_stations[nowrow][12])

    field2 = driver.find_element(By.ID,"eki2_in")
    field2.send_keys(list_stations[nowrow][21])

    submit_button =  driver.find_element(By.CLASS_NAME, "btn.search")
    submit_button.click()
    time.sleep(5)

def extract_commuting(url):
    request = req.Request(url, headers={"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36"})
    with req.urlopen(request) as response:
        data = response.read().decode("utf-8")
    root = bs4.BeautifulSoup(data, "html.parser")



    try:
        fees = root.find("div", {"id": "bR1"}).find("div", class_="tk_total").find("div", class_="_fare")
    except:
        print('error. cant extract html')

    fee = None
    try: 
        if fees:
            fees = fees.b.string.strip()
            fees = fees.replace(",","").replace("円","")
            return fees
        else:
            return None
    except:
        print('error.cant strip fee')

 

if __name__ == '__main__':

    wb = openpyxl.load_workbook("Route_240424.xlsx")
    ws = wb.active
    s1 = wb['Sheet1']
    list_stations = get_values(s1)

    driver = driver()
    for nowrow in range(366):
        if s1.cell(nowrow+2, 35).value ==None:
            driver.get("https://www.jorudan.co.jp/norikae/cgi/nori.cgi?eki1="+list_stations[nowrow+1][12]+"&eki2="+list_stations[nowrow+1][29]+"&via_on=1&eki3="+list_stations[nowrow+1][15]+"&eki4="+list_stations[nowrow+1][18]+"&eki5="+""+"&eki6=&Dyy=2024&Dmm=5&Ddd=1&Dhh=15&Dmn1=0&Dmn2=5&Cfp=1&Czu=2&C7=1&C2=0&C3=0&C1=0&sort=rec&C4=5&C5=0&C6=2&S=%E6%A4%9C%E7%B4%A2&Cmap1=&rf=nr&pg=20&eok1=&eok2=&eok3=&eok4=&eok5=&eok6=&Csg=1")
            time.sleep(10)

            # print("nowrow: ",nowrow)
            # print("datarow:", datarow)       
            # fill_in_data(list_stations, datarow)
            fees = extract_commuting(driver.current_url)
            if fees == None:
                driver.get("https://www.jorudan.co.jp/norikae/cgi/nori.cgi?Sok="+"決+定"+"&eki1="+list_stations[nowrow+1][12]+"&eok1=R-&eki2="+list_stations[nowrow+1][29]+"&eok2=R-&eki3="+list_stations[nowrow+1][15]+"&eok3=R-&eki4="+list_stations[nowrow+1][18]+"&eok4=&eki5=&eok5=&eki6=&eok6=&rf=nr&pg=20&Dyy=2024&Dmm=5&Ddd=1&Dhh=15&Dmn=5&Cway=0&C1=0&C2=0&C3=0&C4=5&C5=0&C6=2&Cmap1=&usm=0&Cway=0&vstp=0&C7=1&via_on=1&Cfp=1&Czu=2&tks=0")
                fees = extract_commuting(driver.current_url)
            print(fees)
            s1.cell(nowrow+2, 35).value = fees
            wb.save("Route_240424.xlsx")

            nowrow += 1
            print("nowrow: ",nowrow)
        # print("datarow:", datarow)

